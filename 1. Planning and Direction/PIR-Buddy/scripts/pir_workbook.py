#!/usr/bin/env python3
"""
pir_workbook.py — read/write helper for the CTI in a Box
"IR-PIR-SIR-and-Intelligence-Collection-Plan" workbook.

Used by the PIR Buddy skill so a CTI analyst never has to hand-edit cells.
It resolves the next ID for you, keeps the IR -> PIR -> SIR lineage consistent,
and validates referential integrity.

SAFETY: this script refuses to write to a file whose name contains "TEMPLATE"
(case-insensitive) unless you pass --force. Recommended flow:

    # 1. make a working copy of the shipped template (once)
    python pir_workbook.py copy \
        -w "IR-PIR-SIR-and-Intelligence-Collection-Plan-TEMPLATE-and-Starting-Examples.xlsx" \
        --out "IR-PIR-SIR-Working-Copy.xlsx"

    # 2. from then on, operate on the working copy in place
    python pir_workbook.py list-irs -w IR-PIR-SIR-Working-Copy.xlsx
    python pir_workbook.py add-pir -w IR-PIR-SIR-Working-Copy.xlsx \
        --ir IR-13 --question "What ransomware groups target our sector?" \
        --source "Mandiant, H-ISAC" --frequency Quarterly \
        --report-type Template --report-method "Quarterly Leadership Report"
    python pir_workbook.py validate -w IR-PIR-SIR-Working-Copy.xlsx

Requires: openpyxl  (pip install openpyxl)
"""

import argparse
import re
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Sheet + column layout (1-based column indices). Data starts on row 2.
# ---------------------------------------------------------------------------
IRS, PIRS, SIRS, PRODUCTS, ICP = "IRs", "PIRs", "SIRs", "Products", "ICP by Source"
DATA_START = 2

COLS = {
    IRS:      {"stakeholder": 1, "ir": 2, "text": 3, "cadence": 4, "decision": 5},
    PIRS:     {"ir": 1, "irtext": 2, "pir": 3, "question": 4, "source": 5,
               "frequency": 6, "rtype": 7, "rmethod": 8},
    SIRS:     {"ir": 1, "irtext": 2, "pir": 3, "pirtext": 4, "sir": 5, "sirtext": 6,
               "source": 7, "frequency": 8, "rtype": 9, "rmethod": 10},
    PRODUCTS: {"ir": 1, "irtext": 2, "pir": 3, "pirtext": 4, "first_product": 5},
    ICP:      {"source": 1, "frequency": 2, "pir": 3, "desc": 4, "time": 5, "perweek": 6},
}

IR_RE  = re.compile(r"^IR-(\d+)$", re.I)
PIR_RE = re.compile(r"^PIR\.(\d+)\.(\d+)$", re.I)
SIR_RE = re.compile(r"^PIR\.(\d+)\.(\d+)\.SIR\.(\d+)$", re.I)


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------
def load(path):
    p = Path(path)
    if not p.exists():
        sys.exit(f"Workbook not found: {path}")
    try:
        return openpyxl.load_workbook(p)
    except Exception as exc:  # noqa: BLE001
        sys.exit(f"Could not open workbook: {exc}")


def sheet(wb, name):
    if name not in wb.sheetnames:
        sys.exit(f"Sheet '{name}' not found. Sheets: {wb.sheetnames}")
    return wb[name]


def cell(ws, row, key, sheetname):
    return ws.cell(row=row, column=COLS[sheetname][key]).value


def s(v):
    return "" if v is None else str(v).strip()


def rows_where(ws, sheetname, key_col, predicate=None):
    """Yield (row_index, value_of_key_col) for populated data rows."""
    col = COLS[sheetname][key_col]
    for r in range(DATA_START, ws.max_row + 1):
        val = ws.cell(row=r, column=col).value
        if s(val) == "":
            continue
        if predicate is None or predicate(val):
            yield r, val


def last_data_row(ws, sheetname, key_col):
    last = DATA_START - 1
    for r, _ in rows_where(ws, sheetname, key_col):
        last = r
    return last


def is_ref_row(text):
    """'see above' shorthand rows are references, not real requirements."""
    return "see above" in s(text).lower()


# ---------------------------------------------------------------------------
# ID lookups / generation
# ---------------------------------------------------------------------------
def ir_number(ir_id):
    m = IR_RE.match(s(ir_id))
    if not m:
        sys.exit(f"Malformed IR id '{ir_id}' (expected IR-<n>)")
    return int(m.group(1))


def find_ir(ws, ir_id, stakeholder=None):
    matches = []
    for r, val in rows_where(ws, IRS, "ir"):
        if s(val).lower() == s(ir_id).lower() and not is_ref_row(cell(ws, r, "text", IRS)):
            if stakeholder is None or s(cell(ws, r, "stakeholder", IRS)).lower() == stakeholder.lower():
                matches.append(r)
    return matches


def find_pir(ws, pir_id):
    for r, val in rows_where(ws, PIRS, "pir"):
        if s(val).lower() == s(pir_id).lower():
            return r
    return None


def find_sir(ws, sir_id):
    for r, val in rows_where(ws, SIRS, "sir"):
        if s(val).lower() == s(sir_id).lower():
            return r
    return None


def next_ir_id(ws):
    mx = 0
    for _, val in rows_where(ws, IRS, "ir"):
        m = IR_RE.match(s(val))
        if m:
            mx = max(mx, int(m.group(1)))
    return f"IR-{mx + 1}"


def next_pir_id(ws, ir_id):
    n = ir_number(ir_id)
    mx = 0
    for _, val in rows_where(ws, PIRS, "pir"):
        m = PIR_RE.match(s(val))
        if m and int(m.group(1)) == n:
            mx = max(mx, int(m.group(2)))
    return f"PIR.{n}.{mx + 1}"


def next_sir_id(ws, pir_id):
    m = PIR_RE.match(s(pir_id))
    if not m:
        sys.exit(f"Malformed PIR id '{pir_id}' (expected PIR.<n>.<m>)")
    a, b = int(m.group(1)), int(m.group(2))
    mx = 0
    for _, val in rows_where(ws, SIRS, "sir"):
        sm = SIR_RE.match(s(val))
        if sm and int(sm.group(1)) == a and int(sm.group(2)) == b:
            mx = max(mx, int(sm.group(3)))
    return f"PIR.{a}.{b}.SIR.{mx + 1}"


# ---------------------------------------------------------------------------
# Save (with template guard)
# ---------------------------------------------------------------------------
def save(wb, workbook_path, out, force, what):
    target = Path(out) if out else Path(workbook_path)
    if "template" in target.name.lower() and not force:
        sys.exit(
            f"Refusing to write to '{target.name}' — it looks like the shipped TEMPLATE.\n"
            f"Make a working copy first:\n"
            f'    python pir_workbook.py copy -w "{workbook_path}" --out "IR-PIR-SIR-Working-Copy.xlsx"\n'
            f"then re-run against the working copy (or pass --force to override)."
        )
    wb.save(target)
    print(f"OK: {what}")
    print(f"    saved -> {target}")


# ---------------------------------------------------------------------------
# Commands: read
# ---------------------------------------------------------------------------
def cmd_copy(a):
    src, dst = Path(a.workbook), Path(a.out)
    if not src.exists():
        sys.exit(f"Source workbook not found: {src}")
    if dst.exists() and not a.force:
        sys.exit(f"'{dst}' already exists. Pass --force to overwrite.")
    shutil.copyfile(src, dst)
    print(f"OK: copied working workbook\n    {src}\n    -> {dst}")


def cmd_list_stakeholders(a):
    ws = sheet(load(a.workbook), IRS)
    seen = {}
    for r, val in rows_where(ws, IRS, "stakeholder"):
        if is_ref_row(cell(ws, r, "text", IRS)):
            continue
        seen[s(val)] = seen.get(s(val), 0) + 1
    if not seen:
        print("No stakeholders found.")
        return
    print(f"{'Stakeholder':<52} {'# IR rows'}")
    print("-" * 64)
    for name in sorted(seen):
        print(f"{name:<52} {seen[name]}")


def cmd_list_irs(a):
    ws = sheet(load(a.workbook), IRS)
    printed = 0
    for r, val in rows_where(ws, IRS, "ir"):
        stk = s(cell(ws, r, "stakeholder", IRS))
        txt = s(cell(ws, r, "text", IRS))
        if is_ref_row(txt):
            continue
        if a.stakeholder and a.stakeholder.lower() not in stk.lower():
            continue
        if a.ir and s(val).lower() != a.ir.lower():
            continue
        dec = s(cell(ws, r, "decision", IRS))
        cad = s(cell(ws, r, "cadence", IRS))
        print(f"[{s(val)}] ({stk} | {cad})")
        print(f"    Q: {txt}")
        print(f"    Decision: {dec}")
        printed += 1
    if not printed:
        print("No matching IRs.")


def cmd_list_pirs(a):
    ws = sheet(load(a.workbook), PIRS)
    printed = 0
    for r, val in rows_where(ws, PIRS, "pir"):
        ir = s(cell(ws, r, "ir", PIRS))
        if a.ir and ir.lower() != a.ir.lower():
            continue
        print(f"[{s(val)}] (parent {ir})")
        print(f"    Q: {s(cell(ws, r, 'question', PIRS))}")
        print(f"    Source: {s(cell(ws, r, 'source', PIRS))} | "
              f"Freq: {s(cell(ws, r, 'frequency', PIRS))} | "
              f"Report: {s(cell(ws, r, 'rtype', PIRS))} -> {s(cell(ws, r, 'rmethod', PIRS))}")
        printed += 1
    if not printed:
        print("No matching PIRs.")


def cmd_list_sirs(a):
    ws = sheet(load(a.workbook), SIRS)
    printed = 0
    for r, val in rows_where(ws, SIRS, "sir"):
        pir = s(cell(ws, r, "pir", SIRS))
        if a.pir and pir.lower() != a.pir.lower():
            continue
        print(f"[{s(val)}] (parent {pir})")
        print(f"    {s(cell(ws, r, 'sirtext', SIRS))}")
        print(f"    Source: {s(cell(ws, r, 'source', SIRS))} | "
              f"Freq: {s(cell(ws, r, 'frequency', SIRS))}")
        printed += 1
    if not printed:
        print("No matching SIRs.")


def cmd_list_products(a):
    ws = sheet(load(a.workbook), PRODUCTS)
    first = COLS[PRODUCTS]["first_product"]
    cols = []
    for c in range(first, ws.max_column + 1):
        h = s(ws.cell(row=1, column=c).value)
        if h:
            cols.append((c, h))
    if not cols:
        print("No product columns found.")
        return
    print("Reporting products (columns on the Products sheet):")
    for _, h in cols:
        print(f"  - {h}")


def cmd_next_id(a):
    wb = load(a.workbook)
    if a.type == "ir":
        print(next_ir_id(sheet(wb, IRS)))
    elif a.type == "pir":
        if not a.parent:
            sys.exit("--parent IR-<n> is required for --type pir")
        print(next_pir_id(sheet(wb, PIRS), a.parent))
    elif a.type == "sir":
        if not a.parent:
            sys.exit("--parent PIR.<n>.<m> is required for --type sir")
        print(next_sir_id(sheet(wb, SIRS), a.parent))


def cmd_show(a):
    wb = load(a.workbook)
    ident = a.id
    if IR_RE.match(ident):
        ws = sheet(wb, IRS)
        rows = find_ir(ws, ident)
        if not rows:
            print(f"{ident} not found.")
            return
        for r in rows:
            print(f"[{ident}] {s(cell(ws, r, 'stakeholder', IRS))}")
            print(f"    Q: {s(cell(ws, r, 'text', IRS))}")
            print(f"    Cadence: {s(cell(ws, r, 'cadence', IRS))}")
            print(f"    Decision: {s(cell(ws, r, 'decision', IRS))}")
    elif PIR_RE.match(ident):
        ws = sheet(wb, PIRS)
        r = find_pir(ws, ident)
        if not r:
            print(f"{ident} not found.")
            return
        for k in ("ir", "question", "source", "frequency", "rtype", "rmethod"):
            print(f"    {k}: {s(cell(ws, r, k, PIRS))}")
    elif SIR_RE.match(ident):
        ws = sheet(wb, SIRS)
        r = find_sir(ws, ident)
        if not r:
            print(f"{ident} not found.")
            return
        for k in ("ir", "pir", "sirtext", "source", "frequency", "rtype", "rmethod"):
            print(f"    {k}: {s(cell(ws, r, k, SIRS))}")
    else:
        sys.exit(f"Unrecognized id '{ident}' (expected IR-n, PIR.n.m, or PIR.n.m.SIR.k)")


# ---------------------------------------------------------------------------
# Commands: write
# ---------------------------------------------------------------------------
def cmd_add_ir(a):
    wb = load(a.workbook)
    ws = sheet(wb, IRS)
    ir_id = a.ir or next_ir_id(ws)
    if not IR_RE.match(ir_id):
        sys.exit(f"--ir must look like IR-<n>, got '{ir_id}'")
    row = last_data_row(ws, IRS, "ir") + 1
    # append after the true last used row across the sheet
    row = max(row, ws.max_row + 1)
    ws.cell(row=row, column=COLS[IRS]["stakeholder"], value=a.stakeholder)
    ws.cell(row=row, column=COLS[IRS]["ir"], value=ir_id)
    ws.cell(row=row, column=COLS[IRS]["text"], value=a.text)
    ws.cell(row=row, column=COLS[IRS]["cadence"], value=a.cadence or "")
    ws.cell(row=row, column=COLS[IRS]["decision"], value=a.decision or "")
    if not s(a.decision):
        print("WARN: no --decision given. An IR without a 'so what' decision is not a real requirement.")
    save(wb, a.workbook, a.out, a.force, f"added {ir_id} for '{a.stakeholder}'")


def cmd_add_pir(a):
    wb = load(a.workbook)
    ws_ir, ws = sheet(wb, IRS), sheet(wb, PIRS)
    ir_rows = find_ir(ws_ir, a.ir)
    if not ir_rows and not a.force:
        sys.exit(f"Parent {a.ir} not found on IRs sheet. Add the IR first, or pass --force.")
    irtext = s(cell(ws_ir, ir_rows[0], "text", IRS)) if ir_rows else ""
    pir_id = a.pir or next_pir_id(ws, a.ir)
    if not PIR_RE.match(pir_id):
        sys.exit(f"--pir must look like PIR.<n>.<m>, got '{pir_id}'")
    if find_pir(ws, pir_id):
        sys.exit(f"{pir_id} already exists. Use 'modify' to change it or omit --pir to auto-number.")
    row = ws.max_row + 1
    ws.cell(row=row, column=COLS[PIRS]["ir"], value=a.ir)
    ws.cell(row=row, column=COLS[PIRS]["irtext"], value=irtext)
    ws.cell(row=row, column=COLS[PIRS]["pir"], value=pir_id)
    ws.cell(row=row, column=COLS[PIRS]["question"], value=a.question)
    ws.cell(row=row, column=COLS[PIRS]["source"], value=a.source or "")
    ws.cell(row=row, column=COLS[PIRS]["frequency"], value=a.frequency or "")
    ws.cell(row=row, column=COLS[PIRS]["rtype"], value=a.report_type or "")
    ws.cell(row=row, column=COLS[PIRS]["rmethod"], value=a.report_method or "")
    if not s(a.source):
        print("WARN: no --source given. A PIR should name at least one collection source.")
    save(wb, a.workbook, a.out, a.force, f"added {pir_id} under {a.ir}")


def cmd_add_sir(a):
    wb = load(a.workbook)
    ws_pir, ws = sheet(wb, PIRS), sheet(wb, SIRS)
    pr = find_pir(ws_pir, a.pir)
    if pr is None and not a.force:
        sys.exit(f"Parent PIR {a.pir} not found on PIRs sheet. Add the PIR first, or pass --force.")
    ir = s(cell(ws_pir, pr, "ir", PIRS)) if pr else ""
    irtext = s(cell(ws_pir, pr, "irtext", PIRS)) if pr else ""
    pirtext = s(cell(ws_pir, pr, "question", PIRS)) if pr else ""
    sir_id = a.sir or next_sir_id(ws, a.pir)
    if not SIR_RE.match(sir_id):
        sys.exit(f"--sir must look like PIR.<n>.<m>.SIR.<k>, got '{sir_id}'")
    if find_sir(ws, sir_id):
        sys.exit(f"{sir_id} already exists.")
    row = ws.max_row + 1
    m = {"ir": ir, "irtext": irtext, "pir": a.pir, "pirtext": pirtext, "sir": sir_id,
         "sirtext": a.text, "source": a.source or "", "frequency": a.frequency or "",
         "rtype": a.report_type or "", "rmethod": a.report_method or ""}
    for k, v in m.items():
        ws.cell(row=row, column=COLS[SIRS][k], value=v)
    save(wb, a.workbook, a.out, a.force, f"added {sir_id} under {a.pir}")


def _product_column(ws, name):
    first = COLS[PRODUCTS]["first_product"]
    for c in range(first, ws.max_column + 1):
        if s(ws.cell(row=1, column=c).value).lower() == name.lower():
            return c
    # partial match fallback
    for c in range(first, ws.max_column + 1):
        h = s(ws.cell(row=1, column=c).value).lower()
        if h and name.lower() in h:
            return c
    return None


def cmd_add_product_row(a):
    wb = load(a.workbook)
    ws_pir, ws = sheet(wb, PIRS), sheet(wb, PRODUCTS)
    pr = find_pir(ws_pir, a.pir)
    if pr is None and not a.force:
        sys.exit(f"PIR {a.pir} not found on PIRs sheet (pass --force to add anyway).")
    ir = s(cell(ws_pir, pr, "ir", PIRS)) if pr else (a.ir or "")
    irtext = s(cell(ws_pir, pr, "irtext", PIRS)) if pr else ""
    pirtext = s(cell(ws_pir, pr, "question", PIRS)) if pr else ""
    row = ws.max_row + 1
    ws.cell(row=row, column=COLS[PRODUCTS]["ir"], value=ir)
    ws.cell(row=row, column=COLS[PRODUCTS]["irtext"], value=irtext)
    ws.cell(row=row, column=COLS[PRODUCTS]["pir"], value=a.pir)
    ws.cell(row=row, column=COLS[PRODUCTS]["pirtext"], value=pirtext)
    for pair in a.set or []:
        if "=" not in pair:
            sys.exit(f"--set expects Product=Yes|No, got '{pair}'")
        pname, yn = pair.split("=", 1)
        col = _product_column(ws, pname.strip())
        if col is None:
            avail = [s(ws.cell(row=1, column=c).value) for c in
                     range(COLS[PRODUCTS]["first_product"], ws.max_column + 1)
                     if s(ws.cell(row=1, column=c).value)]
            sys.exit(f"Product column '{pname.strip()}' not found. Available: {avail}")
        ws.cell(row=row, column=col, value=yn.strip())
    save(wb, a.workbook, a.out, a.force, f"added Products row for {a.pir}")


def cmd_add_icp_row(a):
    wb = load(a.workbook)
    ws_pir, ws = sheet(wb, PIRS), sheet(wb, ICP)
    pr = find_pir(ws_pir, a.pir)
    if pr is None and not a.force:
        sys.exit(f"PIR {a.pir} not found on PIRs sheet (pass --force to add anyway).")
    desc = a.desc or (s(cell(ws_pir, pr, "question", PIRS)) if pr else "")
    row = ws.max_row + 1
    ws.cell(row=row, column=COLS[ICP]["source"], value=a.source)
    ws.cell(row=row, column=COLS[ICP]["frequency"], value=a.frequency or "")
    ws.cell(row=row, column=COLS[ICP]["pir"], value=a.pir)
    ws.cell(row=row, column=COLS[ICP]["desc"], value=desc)
    ws.cell(row=row, column=COLS[ICP]["time"], value=a.time or "")
    ws.cell(row=row, column=COLS[ICP]["perweek"], value=a.per_week or "")
    save(wb, a.workbook, a.out, a.force, f"added ICP line: {a.source} x {a.pir}")


# field aliases accepted by `modify --set field=value`
MODIFY_FIELDS = {
    "pir": {"question": "question", "source": "source", "frequency": "frequency",
            "report-type": "rtype", "rtype": "rtype",
            "report-method": "rmethod", "rmethod": "rmethod"},
    "sir": {"text": "sirtext", "sir": "sirtext", "source": "source",
            "frequency": "frequency", "report-type": "rtype", "rtype": "rtype",
            "report-method": "rmethod", "rmethod": "rmethod"},
    "ir": {"stakeholder": "stakeholder", "text": "text", "requirement": "text",
           "question": "text", "cadence": "cadence", "decision": "decision"},
}


def cmd_modify(a):
    wb = load(a.workbook)
    ident = a.id
    if PIR_RE.match(ident):
        ws, sheetname, fields = sheet(wb, PIRS), PIRS, MODIFY_FIELDS["pir"]
        r = find_pir(ws, ident)
    elif SIR_RE.match(ident):
        ws, sheetname, fields = sheet(wb, SIRS), SIRS, MODIFY_FIELDS["sir"]
        r = find_sir(ws, ident)
    elif IR_RE.match(ident):
        ws, sheetname, fields = sheet(wb, IRS), IRS, MODIFY_FIELDS["ir"]
        matches = find_ir(ws, ident, a.stakeholder)
        if len(matches) > 1:
            stks = [s(cell(ws, m, "stakeholder", IRS)) for m in matches]
            sys.exit(f"{ident} appears for multiple stakeholders {stks}. "
                     f"Disambiguate with --stakeholder.")
        r = matches[0] if matches else None
    else:
        sys.exit(f"Unrecognized id '{ident}'.")
    if not r:
        sys.exit(f"{ident} not found.")
    changed = []
    for pair in a.set or []:
        if "=" not in pair:
            sys.exit(f"--set expects field=value, got '{pair}'")
        f, v = pair.split("=", 1)
        key = fields.get(f.strip().lower())
        if not key:
            sys.exit(f"Field '{f.strip()}' not editable on {sheetname}. Allowed: {sorted(set(fields))}")
        ws.cell(row=r, column=COLS[sheetname][key], value=v.strip())
        changed.append(f"{f.strip()} -> '{v.strip()}'")
    if not changed:
        sys.exit("Nothing to change. Pass one or more --set field=value.")
    save(wb, a.workbook, a.out, a.force, f"modified {ident}: " + "; ".join(changed))


# ---------------------------------------------------------------------------
# Command: validate
# ---------------------------------------------------------------------------
def cmd_validate(a):
    wb = load(a.workbook)
    ws_ir, ws_pir, ws_sir, ws_icp = (sheet(wb, IRS), sheet(wb, PIRS),
                                     sheet(wb, SIRS), sheet(wb, ICP))
    errors, warnings = [], []

    ir_ids = set()
    for r, val in rows_where(ws_ir, IRS, "ir"):
        if is_ref_row(cell(ws_ir, r, "text", IRS)):
            continue
        vid = s(val)
        if not IR_RE.match(vid):
            errors.append(f"IRs row {r}: malformed IR id '{vid}'")
        ir_ids.add(vid.upper())
        if not s(cell(ws_ir, r, "decision", IRS)):
            warnings.append(f"IRs row {r} ({vid}): blank Decision (no 'so what').")

    pir_ids = {}
    for r, val in rows_where(ws_pir, PIRS, "pir"):
        vid = s(val)
        m = PIR_RE.match(vid)
        if not m:
            errors.append(f"PIRs row {r}: malformed PIR id '{vid}'")
            continue
        if vid.upper() in pir_ids:
            errors.append(f"PIRs row {r}: duplicate PIR id '{vid}'")
        pir_ids[vid.upper()] = r
        parent = s(cell(ws_pir, r, "ir", PIRS))
        if parent.upper() not in ir_ids:
            errors.append(f"PIRs row {r} ({vid}): parent {parent} not found on IRs sheet.")
        elif ir_number(parent) != int(m.group(1)):
            errors.append(f"PIRs row {r} ({vid}): number prefix does not match parent {parent}.")
        if not s(cell(ws_pir, r, "source", PIRS)):
            warnings.append(f"PIRs row {r} ({vid}): no Collection Source.")

    seen_sir = set()
    for r, val in rows_where(ws_sir, SIRS, "sir"):
        vid = s(val)
        m = SIR_RE.match(vid)
        if not m:
            errors.append(f"SIRs row {r}: malformed SIR id '{vid}'")
            continue
        if vid.upper() in seen_sir:
            errors.append(f"SIRs row {r}: duplicate SIR id '{vid}'")
        seen_sir.add(vid.upper())
        parent = s(cell(ws_sir, r, "pir", SIRS))
        if parent.upper() not in pir_ids:
            errors.append(f"SIRs row {r} ({vid}): parent PIR {parent} not found on PIRs sheet.")

    for r, val in rows_where(ws_icp, ICP, "pir"):
        vid = s(val)
        if PIR_RE.match(vid) and vid.upper() not in pir_ids:
            warnings.append(f"ICP by Source row {r}: PIR {vid} not found on PIRs sheet.")

    print(f"IRs: {len(ir_ids)} unique · PIRs: {len(pir_ids)} · SIRs: {len(seen_sir)}")
    for w in warnings:
        print(f"  WARN  {w}")
    for e in errors:
        print(f"  ERROR {e}")
    if errors:
        print(f"\nVALIDATION FAILED — {len(errors)} error(s), {len(warnings)} warning(s).")
        sys.exit(1)
    print(f"\nVALIDATION PASSED — 0 errors, {len(warnings)} warning(s).")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def build_parser():
    p = argparse.ArgumentParser(description="CTI IR/PIR/SIR/ICP workbook helper.")
    sub = p.add_subparsers(dest="cmd", required=True)

    def add_wb(sp, out=False):
        sp.add_argument("-w", "--workbook", required=True, help="path to the .xlsx workbook")
        if out:
            sp.add_argument("--out", help="write to this file instead of in place")
            sp.add_argument("--force", action="store_true",
                            help="allow writing to a TEMPLATE-named file / skip parent checks")

    sp = sub.add_parser("copy", help="copy the template to a working workbook")
    sp.add_argument("-w", "--workbook", required=True, help="source workbook")
    sp.add_argument("--out", required=True, help="destination working workbook")
    sp.add_argument("--force", action="store_true", help="overwrite destination if it exists")
    sp.set_defaults(func=cmd_copy)

    sp = sub.add_parser("list-stakeholders"); add_wb(sp); sp.set_defaults(func=cmd_list_stakeholders)

    sp = sub.add_parser("list-irs"); add_wb(sp)
    sp.add_argument("--stakeholder"); sp.add_argument("--ir")
    sp.set_defaults(func=cmd_list_irs)

    sp = sub.add_parser("list-pirs"); add_wb(sp)
    sp.add_argument("--ir"); sp.set_defaults(func=cmd_list_pirs)

    sp = sub.add_parser("list-sirs"); add_wb(sp)
    sp.add_argument("--pir"); sp.set_defaults(func=cmd_list_sirs)

    sp = sub.add_parser("list-products"); add_wb(sp); sp.set_defaults(func=cmd_list_products)

    sp = sub.add_parser("show"); add_wb(sp)
    sp.add_argument("id"); sp.set_defaults(func=cmd_show)

    sp = sub.add_parser("next-id"); add_wb(sp)
    sp.add_argument("--type", required=True, choices=["ir", "pir", "sir"])
    sp.add_argument("--parent"); sp.set_defaults(func=cmd_next_id)

    sp = sub.add_parser("add-ir"); add_wb(sp, out=True)
    sp.add_argument("--stakeholder", required=True)
    sp.add_argument("--text", required=True, help="the intelligence requirement question")
    sp.add_argument("--decision", help="the 'so what' action this IR drives")
    sp.add_argument("--cadence"); sp.add_argument("--ir", help="force a specific IR-<n>")
    sp.set_defaults(func=cmd_add_ir)

    sp = sub.add_parser("add-pir"); add_wb(sp, out=True)
    sp.add_argument("--ir", required=True, help="parent IR-<n>")
    sp.add_argument("--question", required=True)
    sp.add_argument("--source"); sp.add_argument("--frequency")
    sp.add_argument("--report-type"); sp.add_argument("--report-method")
    sp.add_argument("--pir", help="force a specific PIR.<n>.<m>")
    sp.set_defaults(func=cmd_add_pir)

    sp = sub.add_parser("add-sir"); add_wb(sp, out=True)
    sp.add_argument("--pir", required=True, help="parent PIR.<n>.<m>")
    sp.add_argument("--text", required=True, help="the specific question (named actor/vendor/CVE)")
    sp.add_argument("--source"); sp.add_argument("--frequency")
    sp.add_argument("--report-type"); sp.add_argument("--report-method")
    sp.add_argument("--sir", help="force a specific SIR id")
    sp.set_defaults(func=cmd_add_sir)

    sp = sub.add_parser("add-product-row"); add_wb(sp, out=True)
    sp.add_argument("--pir", required=True)
    sp.add_argument("--ir", help="used only if the PIR is not found and --force is set")
    sp.add_argument("--set", action="append", metavar="Product=Yes|No",
                    help="repeatable; e.g. --set 'Monthly CTI Report=Yes'")
    sp.set_defaults(func=cmd_add_product_row)

    sp = sub.add_parser("add-icp-row"); add_wb(sp, out=True)
    sp.add_argument("--source", required=True)
    sp.add_argument("--pir", required=True)
    sp.add_argument("--frequency"); sp.add_argument("--desc")
    sp.add_argument("--time", help="average collection time, e.g. '3 min'")
    sp.add_argument("--per-week", help="frequency per week (number)")
    sp.set_defaults(func=cmd_add_icp_row)

    sp = sub.add_parser("modify"); add_wb(sp, out=True)
    sp.add_argument("--id", required=True, help="IR-n, PIR.n.m, or PIR.n.m.SIR.k")
    sp.add_argument("--stakeholder", help="disambiguate when modifying a reused IR")
    sp.add_argument("--set", action="append", metavar="field=value",
                    help="repeatable; e.g. --set 'source=Mandiant, H-ISAC'")
    sp.set_defaults(func=cmd_modify)

    sp = sub.add_parser("validate"); add_wb(sp); sp.set_defaults(func=cmd_validate)
    return p


def main(argv=None):
    args = build_parser().parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
